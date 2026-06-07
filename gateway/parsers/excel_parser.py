"""Excel (.xlsx, .xls) → unified JSON. Uses openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from gateway.parsers.common import ParserError, summarize_rows


def parse_excel(path: Path) -> dict[str, Any]:
    """Parse the FIRST sheet of an Excel file into the unified format.

    Multi-sheet handling is a Phase 2+ concern; for v1 we only ship the
    first sheet (most data-analysis report scenarios are single-sheet).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ParserError("openpyxl not installed — cannot parse Excel files") from e
    if not path.is_file():
        raise ParserError(f"Excel file not found: {path}")
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:  # openpyxl raises a wide variety of exceptions
        raise ParserError(f"Cannot open Excel file: {e}") from e
    try:
        sheets = wb.sheetnames
        if not sheets:
            raise ParserError("Excel file has no sheets")
        ws = wb[sheets[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {
                "format": "excel",
                "rows": [],
                "summary": {"row_count": 0, "columns": [], "sheets": sheets, "active_sheet": sheets[0]},
            }
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        rows: list[dict[str, Any]] = []
        for r in rows_iter:
            if all(c is None for c in r):
                continue  # skip fully-empty rows
            row = {}
            for h, v in zip(headers, r):
                if isinstance(v, (int, float, str, bool)) or v is None:
                    row[h] = v
                else:
                    # datetime, etc → stringify
                    row[h] = str(v)
            rows.append(row)
    finally:
        wb.close()
    summary = summarize_rows(rows)
    summary["sheets"] = sheets
    summary["active_sheet"] = sheets[0]
    return {
        "format": "excel",
        "rows": rows,
        "summary": summary,
    }
